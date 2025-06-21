from openpyxl import load_workbook

file_path = "data.xlsx"  
wb = load_workbook(file_path, data_only=True)

sheet = wb["Page_0"]

headers = [cell.value for cell in sheet[3]]
product_index = headers.index('Product')
price_index = headers.index('Price')

total_price = 0
count = 0

for row in sheet.iter_rows(min_row=4, values_only=True):
    product = row[product_index]
    price = row[price_index]

    if isinstance(product, str) and "LaserJet" in product:
        try:
            price_val = float(price)
            total_price += price_val
            count += 1
        except (TypeError, ValueError):
            continue

average = int(total_price / count) if count > 0 else 0

print("Count:", average)
