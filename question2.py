from openpyxl import load_workbook
from datetime import datetime

file_path = "data.xlsx"
wb = load_workbook(file_path, data_only=True)

sheet = wb["Page_0"]

headers = [cell.value for cell in sheet[3]]
priority_index = headers.index('Priority')
delivery_date_index = headers.index('Delivery date')

count = 0

for row in sheet.iter_rows(min_row=4, values_only=True):
    priority = row[priority_index]
    delivery_date = row[delivery_date_index]

    if priority == 'High' and isinstance(delivery_date, datetime):
        if delivery_date.year == 2015:
            count += 1

print("Count: ", count)
