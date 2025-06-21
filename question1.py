from openpyxl import load_workbook

file_path = "data.xlsx" 
wb = load_workbook(file_path, data_only=True)

sheet = wb["Page_0"]

headers = [cell.value for cell in sheet[3]]
address_index = headers.index('Address')
quantity_index = headers.index('Quantity')

count = 0

for row in sheet.iter_rows(min_row=4, values_only=True):
    address = row[address_index]
    quantity = row[quantity_index]

    if isinstance(address, str) and address.startswith("Ain"):
        try:
            if float(quantity) < 40:
                count += 1
        except (TypeError, ValueError):
            continue 

print("Count: ", count)
