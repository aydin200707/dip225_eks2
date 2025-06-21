from openpyxl import load_workbook

file_path = "data.xlsx"
wb = load_workbook(file_path, data_only=True)

sheet = wb["Page_0"]

headers = [cell.value for cell in sheet[3]]
address_index = headers.index('Address')
city_index = headers.index('City')

count = 0

for row in sheet.iter_rows(min_row=4, values_only=True):
    address = row[address_index]
    city = row[city_index]

    if isinstance(address, str) and "Adulienas iela" in address:
        if city in ("Valmiera", "Saulkrasti"):
            count += 1

print("Number of matching entries:", count)
