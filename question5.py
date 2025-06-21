from openpyxl import load_workbook

file_path = "sagatave_eksamenam.xlsx" 
wb = load_workbook(file_path, data_only=True)

sheet = wb["Page_0"]

headers = [cell.value for cell in sheet[3]]
client_index = headers.index('Client')
quantity_index = headers.index('Quantity')
total_index = headers.index('Total')

total_sum = 0

for row in sheet.iter_rows(min_row=4, values_only=True):
    client = row[client_index]
    quantity = row[quantity_index]
    total = row[total_index]

    if client == "Korporatīvais":
        try:
            quantity_val = float(quantity)
            total_val = float(total)
            if 40 <= quantity_val <= 50:
                total_sum += total_val
        except (TypeError, ValueError):
            continue

result = int(total_sum)
print("Total sum (rounded down) for Korporatīvais clients with quantity 40–50:", result)
